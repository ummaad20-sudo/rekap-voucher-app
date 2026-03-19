from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.scrollview import ScrollView
from kivy.graphics import Color, Rectangle
from kivy.utils import get_color_from_hex
from kivy.core.clipboard import Clipboard

from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime


class RekapApp(App):

    def build(self):

        self.data_manual = []

        root = BoxLayout(orientation='vertical', padding=15, spacing=10)

        # Background
        with root.canvas.before:
            Color(rgba=get_color_from_hex("#1e272e"))
            self.bg = Rectangle(size=root.size, pos=root.pos)

        root.bind(size=self.update_bg, pos=self.update_bg)

        # TITLE
        title = Label(
            text="[b]REKAP VOUCHER PRO[/b]",
            markup=True,
            size_hint=(1, 0.08),
            font_size=22,
            color=(1, 1, 1, 1)
        )
        root.add_widget(title)

        # =====================
        # INPUT FILE PATH
        # =====================
        self.file_input = TextInput(
            hint_text="Path Excel\nContoh:\n/storage/emulated/0/Download/data.xlsx",
            size_hint=(1, 0.12)
        )
        root.add_widget(self.file_input)

        btn_baca = Button(
            text="BACA FILE EXCEL",
            background_normal="",
            background_color=get_color_from_hex("#e67e22"),
            size_hint=(1, 0.08)
        )
        btn_baca.bind(on_press=self.baca_file)
        root.add_widget(btn_baca)

        # =====================
        # INPUT MANUAL
        # =====================
        self.tanggal = TextInput(
            hint_text="Tanggal (YYYY/MM/DD)",
            size_hint=(1, 0.08)
        )

        self.grup = TextInput(
            hint_text="Grup",
            size_hint=(1, 0.08)
        )

        self.harga = TextInput(
            hint_text="Harga",
            input_filter="int",
            size_hint=(1, 0.08)
        )

        root.add_widget(self.tanggal)
        root.add_widget(self.grup)
        root.add_widget(self.harga)

        btn_tambah = Button(
            text="TAMBAH DATA",
            background_normal="",
            background_color=get_color_from_hex("#27ae60"),
            size_hint=(1, 0.08)
        )
        btn_tambah.bind(on_press=self.tambah_manual)
        root.add_widget(btn_tambah)

        btn_rekap = Button(
            text="REKAP MANUAL",
            background_normal="",
            background_color=get_color_from_hex("#2980b9"),
            size_hint=(1, 0.08)
        )
        btn_rekap.bind(on_press=self.rekap_manual)
        root.add_widget(btn_rekap)

        # =====================
        # HASIL
        # =====================
        self.result = Label(
            text="Menunggu data...",
            markup=True,
            size_hint_y=None,
            color=(1, 1, 1, 1),
            halign="left",
            valign="top"
        )

        self.result.bind(texture_size=self.result.setter('size'))
        self.result.bind(
            width=lambda s, w: s.setter('text_size')(s, (w - 20, None))
        )

        scroll = ScrollView(size_hint=(1, 0.4))
        scroll.add_widget(self.result)
        root.add_widget(scroll)

        # =====================
        # BUTTON BAWAH
        # =====================
        btn_copy = Button(
            text="COPY HASIL",
            background_normal="",
            background_color=get_color_from_hex("#8e44ad"),
            size_hint=(1, 0.07)
        )
        btn_copy.bind(on_press=self.copy_hasil)
        root.add_widget(btn_copy)

        self.notif = Label(
            text="",
            size_hint=(1, 0.05),
            color=(0.9, 0.9, 0.9, 1)
        )
        root.add_widget(self.notif)

        return root

    # =====================
    # BACKGROUND
    # =====================
    def update_bg(self, instance, value):
        self.bg.size = instance.size
        self.bg.pos = instance.pos

    # =====================
    # FORMAT
    # =====================
    def rupiah(self, x):
        return "Rp {:,}".format(int(x)).replace(",", ".")

    # =====================
    # COPY
    # =====================
    def copy_hasil(self, instance):
        Clipboard.copy(self.result.text)
        self.notif.text = "✔ Hasil disalin"

    # =====================
    # BACA EXCEL
    # =====================
    def baca_file(self, instance):

        path = self.file_input.text.strip()

        if not path:
            self.notif.text = "Isi path file dulu!"
            return

        try:
            wb = load_workbook(path)
            sheet = wb.active
        except Exception as e:
            self.result.text = f"[color=ff4444]Gagal buka file\n{e}[/color]"
            return

        self.proses_excel(sheet)

    def proses_excel(self, sheet):

        rekap = defaultdict(lambda: {"jumlah": 0, "total": 0})
        total_harian = defaultdict(int)

        header = [cell.value for cell in sheet[1]]

        try:
            i_grup = header.index("Grup pengguna")
            i_harga = header.index("Harga")
            i_tgl = header.index("Diaktifkan di")
        except:
            self.result.text = "[color=ff4444]Header tidak sesuai[/color]"
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):

            grup = row[i_grup]
            harga = row[i_harga]
            tanggal = row[i_tgl]

            if grup and harga and tanggal:

                if isinstance(tanggal, datetime):
                    tgl = tanggal.strftime("%Y/%m/%d")
                else:
                    tgl = str(tanggal).split(" ")[0]

                try:
                    harga = int(harga)
                except:
                    continue

                key = (tgl, grup)

                rekap[key]["jumlah"] += 1
                rekap[key]["total"] += harga
                total_harian[tgl] += harga

        self.tampilkan_hasil(rekap, total_harian)

    # =====================
    # INPUT MANUAL
    # =====================
    def tambah_manual(self, instance):

        tgl = self.tanggal.text.strip()
        grp = self.grup.text.strip()
        hrg = self.harga.text.strip()

        if not tgl or not grp or not hrg:
            self.notif.text = "Isi semua data!"
            return

        try:
            datetime.strptime(tgl, "%Y/%m/%d")
            hrg = int(hrg)
        except:
            self.notif.text = "Format salah!"
            return

        self.data_manual.append((tgl, grp, hrg))

        self.tanggal.text = ""
        self.grup.text = ""
        self.harga.text = ""

        self.notif.text = "✔ Data ditambahkan"

    def rekap_manual(self, instance):

        if not self.data_manual:
            self.result.text = "Belum ada data"
            return

        rekap = defaultdict(lambda: {"jumlah": 0, "total": 0})
        total_harian = defaultdict(int)

        for tgl, grp, hrg in self.data_manual:
            key = (tgl, grp)
            rekap[key]["jumlah"] += 1
            rekap[key]["total"] += hrg
            total_harian[tgl] += hrg

        self.tampilkan_hasil(rekap, total_harian)

    # =====================
    # TAMPILKAN HASIL
    # =====================
    def tampilkan_hasil(self, rekap, total_harian):

        hasil = "[b]===== HASIL REKAP =====[/b]\n\n"

        sorted_data = sorted(rekap.items())
        last = None

        for (tgl, grp), val in sorted_data:

            if last and tgl != last:
                hasil += f"[color=00ffcc][b]TOTAL : {self.rupiah(total_harian[last])}[/b][/color]\n\n"

            if tgl != last:
                hasil += f"[color=FFA726][b]Tanggal : {tgl}[/b][/color]\n"

            hasil += f"Grup : {grp}\n"
            hasil += f"Jumlah : {val['jumlah']}\n"
            hasil += f"Total : {self.rupiah(val['total'])}\n\n"

            last = tgl

        grand = sum(total_harian.values())
        hasil += f"\n[color=FFD700][b]GRAND TOTAL : {self.rupiah(grand)}[/b][/color]"

        self.result.text = hasil
        self.notif.text = "✔ Rekap selesai"


if __name__ == "__main__":
    RekapApp().run()
